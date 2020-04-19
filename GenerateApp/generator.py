import argparse
import os
import pathlib
import random
import sys
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def generate(args):
    if not os.access(str(args.BasePath), os.F_OK):
        os.mkdir(str(args.BasePath))
    os.chdir(str(args.BasePath))
    for iteration in range(args.ChartsCount):
        random.seed()
        ColumnCount = random.randint(5, 12)
        workbook = xlsxwriter.Workbook('Table' + str(iteration + 1) + '.xlsx')
        worksheet = workbook.add_worksheet()
        LowerBound = random.randint(100, 500)
        UpperBound = random.randint(600, 1000)
        for column in range(ColumnCount):
            value = random.randint(LowerBound, UpperBound)
            worksheet.write(column, 0, column + 1)
            worksheet.write(column, 1, value)
        workbook.close()

        OpenedWb = load_workbook('Table' + str(iteration + 1) + '.xlsx')
        OpenedWS = OpenedWb.get_sheet_by_name('Sheet1')
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 4
        chart1.title = "Bar Chart"
        chart1.y_axis.title = ''
        chart1.x_axis.title = ''

        data = Reference(OpenedWS, min_col=2, min_row=1, max_row=ColumnCount)
        cats = Reference(OpenedWS, min_col=1, min_row=1, max_row=ColumnCount)
        chart1.add_data(data, titles_from_data=False)
        chart1.set_categories(cats)
        chart1.shape = 4
        OpenedWS.add_chart(chart1, "D3")
        OpenedWb.save('Table' + str(iteration + 1) + '.xlsx')



def parse(argv):
    parser = argparse.ArgumentParser(description="""
     Parser of image and tables generator
        """, formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument('ChartsCount', action='store', type=int)
    parser.add_argument('BasePath', action='store', type=pathlib.Path)
    parser.add_argument('BaseName', action='store', type=str)

    return parser.parse_args(argv[1:])


def main():
    args = parse(sys.argv)
    generate(args)


main()
